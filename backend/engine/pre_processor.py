"""
BOQ Pre-Processor Module

This module handles the initial analysis and extraction of BOQ (Bill of Quantities) 
data from Excel files. It identifies potential headers, analyzes worksheet structure,
and extracts BOQ tables with AI assistance.
"""

import pandas as pd
import openpyxl
import time
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import re
import json
import os
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, asdict, field
import logging
from pathlib import Path
import requests
from datetime import datetime
import time

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Suppress pandas warnings about column duplication in console
import warnings
warnings.filterwarnings('ignore', message='DataFrame columns are not unique')

@dataclass
class CellInfo:
    """Detailed cell information"""
    value: Any
    coordinate: str
    is_merged: bool
    merged_range: Optional[str]
    has_formula: bool
    formula: Optional[str]
    is_wrapped: bool
    font_info: Dict
    fill_info: Dict

@dataclass
class WorksheetAnalysis:
    """Comprehensive worksheet analysis"""
    name: str
    dimensions: Tuple[int, int]  # (max_row, max_col)
    frozen_panes: Optional[str]
    merged_cells: List[str]
    suspected_headers: List[Tuple[int, List[str], float]]  # (row, headers, confidence)
    boq_regions: List[Dict]
    data_density: float
    has_wrapped_text: bool
    ai_analysis: Optional[Dict]
    # NEW: Mandatory column mappings for BOQ processing
    column_mappings: Dict = field(default_factory=dict)

@dataclass
class BOQTable:
    """Enhanced BOQ table structure"""
    worksheet_name: str
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    headers: List[str]
    data: pd.DataFrame
    confidence_score: float
    ai_validated: bool
    extraction_notes: List[str]
    cell_details: List[CellInfo]

@dataclass
class FileAnalysis:
    """Complete file analysis"""
    filename: str
    file_size: int
    worksheets: List[WorksheetAnalysis]
    total_boq_tables: int
    extraction_summary: Dict
    processing_time: float

class AdvancedBOQExtractor:
    def __init__(self, gemini_api_key: str = None):
        """Initialize the extractor with Gemini AI integration using direct HTTP requests"""
        
        # Enhanced BOQ patterns with variations
        self.boq_patterns = {
            'sl_no': [
                r'(?i)(sl\.?\s*no\.?|s\.?\s*no\.?|serial\s*no\.?|item\s*no\.?|sr\.?\s*no\.?)',
                r'(?i)(sequence|order|number|#)',
            ],
            'code': [
                r'(?i)(code|item\s*code|work\s*code|activity\s*code)',
                r'(?i)(ref\.?\s*no\.?|reference)',
            ],
            'description': [
                r'(?i)(description|item\s*description|particulars|work\s*description)',
                r'(?i)(activity|task|work\s*item|specification)',
            ],
            'unit': [
                r'(?i)(unit|uom|u\.?o\.?m\.?|units)',
                r'(?i)(measure|measurement)',
            ],
            'quantity': [
                r'(?i)(qty|quantity|cumulative|overall\s*qty|total\s*qty)',
                r'(?i)(volume|count|number)',
            ],
            'rate': [
                r'(?i)(rate|unit\s*rate|net\s*rate|basic\s*rate)',
                r'(?i)(price|cost\s*per\s*unit)',
            ],
            'amount': [
                r'(?i)(amount|total\s*amount|value|total\s*value)',
                r'(?i)(cost|total\s*cost)',
            ],
            'grade': [
                r'(?i)(grade|grade\s*of\s*conc|concrete\s*grade|grade\s*of\s*concrete)',
                r'(?i)(class|type|category)',
            ],
            'cement': [
                r'(?i)(cement\s*co-?eff|cement\s*coefficient|cement|ceff|c\s*eff)',
            ],
            'remarks': [
                r'(?i)(remarks|notes|comments|observations)',
                r'(?i)(specification|remark)',
            ]
        }
        
        # FIXED: Initialize Gemini AI with proper API key handling
        self.api_key = gemini_api_key
        
        # FIXED: Remove the placeholder check that was causing issues
        if not self.api_key or self.api_key.strip() == "":
            logger.warning("Gemini API key not provided. AI features will be disabled.")
            self.ai_enabled = False
        else:
            # Updated API URL to use the correct endpoint and model
            self.base_url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"
            self.cache = {}  # Simple cache to avoid duplicate API calls
            
            # API call counter
            self.api_call_count = 0
            
            self.ai_enabled = True
            logger.info(f"✅ Gemini AI initialized successfully for preprocessor (key length: {len(self.api_key)} chars)")
    
    def validate_excel_file(self, file_path: str) -> bool:
        """Validate if the file is a proper Excel file"""
        try:
            # Check file extension
            valid_extensions = ['.xlsx', '.xls', '.xlsm']
            if not any(file_path.lower().endswith(ext) for ext in valid_extensions):
                logger.warning(f"File {file_path} doesn't have a valid Excel extension")
                return False
            
            # Try to open with openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            wb.close()
            return True
            
        except Exception as e:
            logger.error(f"File validation failed for {file_path}: {str(e)}")
            return False
    
    def analyze_cell_details(self, ws, cell_coord: str) -> CellInfo:
        """Get detailed information about a specific cell"""
        cell = ws[cell_coord]
        
        # Check if cell is part of merged range
        is_merged = False
        merged_range = None
        for merged_cell in ws.merged_cells.ranges:
            if cell.coordinate in merged_cell:
                is_merged = True
                merged_range = str(merged_cell)
                break
        
        # Get font and fill information
        font_info = {
            'name': cell.font.name if cell.font else None,
            'size': cell.font.size if cell.font else None,
            'bold': cell.font.bold if cell.font else False,
            'italic': cell.font.italic if cell.font else False
        }
        
        fill_info = {
            'pattern_type': cell.fill.patternType if cell.fill else None,
            'start_color': str(cell.fill.start_color.rgb) if cell.fill and cell.fill.start_color else None
        }
        
        return CellInfo(
            value=cell_coord,
            coordinate=cell_coord,
            is_merged=is_merged,
            merged_range=merged_range,
            has_formula=cell.data_type == 'f',
            formula=cell.value if cell.data_type == 'f' else None,
            is_wrapped=cell.alignment.wrap_text if cell.alignment else False,
            font_info=font_info,
            fill_info=fill_info
        )
    
    def handle_merged_cells(self, ws, df: pd.DataFrame) -> pd.DataFrame:
        """Handle merged cells by propagating values"""
        logger.info(f"Processing {len(list(ws.merged_cells.ranges))} merged cell ranges")
        
        for merged_range in ws.merged_cells.ranges:
            # Get the top-left cell value
            min_row, min_col, max_row, max_col = merged_range.bounds
            top_left_value = ws.cell(min_row, min_col).value
            
            # Propagate value to all cells in the merged range
            for row in range(min_row - 1, max_row):  # -1 because pandas is 0-indexed
                for col in range(min_col - 1, max_col):
                    if row < len(df) and col < len(df.columns):
                        if pd.isna(df.iloc[row, col]) or df.iloc[row, col] == '':
                            df.iloc[row, col] = top_left_value
        
        return df
    
    def detect_wrapped_text_content(self, ws) -> List[str]:
        """Detect cells with wrapped text and extract their content"""
        wrapped_content = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.alignment and cell.alignment.wrap_text and cell.value:
                    wrapped_content.append(str(cell.value))
        
        logger.info(f"Found {len(wrapped_content)} cells with wrapped text")
        return wrapped_content
    
    def identify_mandatory_columns(self, df: pd.DataFrame, sheet_name: str) -> Dict:
        """Mandate identification of description, unit, and quantity columns"""
        logger.info(f"🔍 MANDATORY COLUMN DETECTION: Analyzing '{sheet_name}' for required BOQ columns")
        
        column_mappings = {
            "description_column": None,
            "unit_column": None, 
            "quantity_column": None,
            "validation_status": "FAILED",
            "error_message": ""
        }
        
        # 1. Find Description Column (MANDATORY) - Enhanced patterns with priority
        desc_patterns_priority = [
            # HIGHEST PRIORITY - exact matches first
            ['description'],
            # High priority - most specific patterns
            ['item description', 'work description', 'job description', 'activity description', 'particulars'],
            # Medium priority - common specific patterns
            ['specification', 'scope of work', 'nature of work', 'work item', 'activity', 'task'],
            # Lower priority - more generic patterns
            ['item', 'items', 'material', 'materials', 'work', 'details', 'desc', 'item name', 'work details', 'scope', 'type of work']
        ]
        
        # Serial number exclusion patterns - these columns should NEVER be description columns
        serial_exclusion_patterns = ['item no', 'sl. no', 'sl.no', 'slno', 's.no', 's no', 'sr. no', 'sr.no', 'srno', 'serial', 'serial no', 'sno', 'item number', 'row number', 'line no', 'entry no']
        
        desc_col = None
        
        # Try patterns in priority order
        for priority_level, patterns in enumerate(desc_patterns_priority):
            for col in df.columns:
                col_lower = str(col).lower().strip()
                
                # SKIP if this is clearly a serial number column
                if any(serial_pattern in col_lower for serial_pattern in serial_exclusion_patterns):
                    continue
                    
                if any(desc_name in col_lower for desc_name in patterns):
                    desc_col = col
                    logger.info(f"✅ DESCRIPTION COLUMN FOUND (Priority {priority_level+1}): '{col}'")
                    break
            if desc_col:
                break
        
        if not desc_col:
            # Fallback: longest text column (excluding serial number columns)
            text_columns = [col for col in df.columns if df[col].dtype == 'object' and 
                          not any(serial_pattern in str(col).lower().strip() for serial_pattern in serial_exclusion_patterns)]
            if text_columns:
                text_lengths = {col: df[col].astype(str).str.len().mean() for col in text_columns}
                desc_col = max(text_lengths, key=text_lengths.get)
                logger.warning(f"⚠️ DESCRIPTION COLUMN (fallback): '{desc_col}'")
        
        if not desc_col:
            column_mappings["error_message"] = "No description column found"
            return column_mappings
        
        column_mappings["description_column"] = desc_col
        
        # 2. Find Unit Column (MANDATORY) - Enhanced patterns with priority
        unit_patterns_priority = [
            # HIGHEST PRIORITY - exact matches first
            ['unit'],
            # High priority - most specific BOQ unit patterns
            ['unit of measurement', 'u.o.m', 'uom', 'measuring unit'],
            # Medium priority - common unit patterns  
            ['units', 'measure', 'measurement', 'unit measure'],
            # Lower priority - abbreviated patterns
            ['un', 'u/m', 'u.m', 'um']
        ]
        unit_col = None
        
        # Try patterns in priority order
        for priority_level, patterns in enumerate(unit_patterns_priority):
            for col in df.columns:
                col_lower = str(col).lower().strip()
                
                # SKIP if this is clearly a serial number column or already identified as description
                if (any(serial_pattern in col_lower for serial_pattern in serial_exclusion_patterns) or 
                    col == desc_col):
                    continue
                    
                if any(unit_name in col_lower for unit_name in patterns):
                    unit_col = col
                    logger.info(f"✅ UNIT COLUMN FOUND (Priority {priority_level+1}): '{col}'")
                    break
            if unit_col:
                break
        
        if not unit_col:
            # Look for short text columns with unit-like values (excluding serial number columns)
            for col in df.columns:
                col_lower = str(col).lower().strip()
                
                if (df[col].dtype == 'object' and col != desc_col and
                    not any(serial_pattern in col_lower for serial_pattern in serial_exclusion_patterns)):
                    avg_length = df[col].astype(str).str.len().mean()
                    if avg_length < 10:  # Short text
                        sample_values = df[col].dropna().astype(str).str.lower().unique()[:10]
                        unit_indicators = ['cum', 'sqm', 'mt', 'kg', 'nos', 'rmt', 'ltr', 'm', 'each', 'nos.', 'sq.m', 'cu.m']
                        if any(any(indicator in val for indicator in unit_indicators) for val in sample_values):
                            unit_col = col
                            logger.info(f"✅ UNIT COLUMN DETECTED (by pattern): '{col}'")
                            break
        
        if not unit_col:
            column_mappings["error_message"] = f"No unit column found in sheet '{sheet_name}'"
            return column_mappings
        
        column_mappings["unit_column"] = unit_col
        
        # 3. Find Quantity Column (MANDATORY) - Enhanced patterns with priority
        qty_patterns_priority = [
            # HIGHEST PRIORITY - exact matches first
            ['quantity'],
            # High priority - most specific quantity patterns
            ['cumulative quantity', 'overall quantity', 'total quantity', 'net quantity', 'gross quantity', 'variation in quantity', 'variation in qty'],
            # Medium priority - common quantity patterns
            ['cumulative', 'overall qty', 'total qty', 'cum qty', 'cumulative qty', 'variation qty', 'var qty', 'variance qty'],
            # Lower priority - abbreviated patterns (removed problematic 'nos' that could match 'nos.' in 'item nos.')
            ['qty', 'qnty', 'quant', 'volume', 'count', 'overall', 'cum', 'qtty', 'amount', 'variation', 'var']
        ]
        qty_col = None
        
        # Try patterns in priority order
        for priority_level, patterns in enumerate(qty_patterns_priority):
            for col in df.columns:
                col_lower = str(col).lower().strip()
                
                # SKIP if this is clearly a serial number column
                if any(serial_pattern in col_lower for serial_pattern in serial_exclusion_patterns):
                    continue
                    
                # SKIP if this is already identified as description or unit column
                if col == desc_col or col == unit_col:
                    continue
                    
                if any(qty_name in col_lower for qty_name in patterns):
                    qty_col = col
                    logger.info(f"✅ QUANTITY COLUMN FOUND (Priority {priority_level+1}): '{col}'")
                    break
            if qty_col:
                break
        
        if not qty_col:
            # Look for numeric columns (excluding serial number columns)
            numeric_cols = df.select_dtypes(include=['int64', 'float64', 'Int64', 'Float64']).columns
            for col in numeric_cols:
                col_lower = str(col).lower().strip()
                
                # SKIP if this is a serial number column, description column, or unit column
                if (col not in [desc_col, unit_col] and 
                    not any(serial_pattern in col_lower for serial_pattern in serial_exclusion_patterns)):
                    # Check if values look like quantities
                    sample_values = df[col].dropna().head(10)
                    if len(sample_values) > 0 and all(val >= 0 for val in sample_values):
                        qty_col = col
                        logger.info(f"✅ QUANTITY COLUMN DETECTED (numeric): '{col}'")
                        break
        
        if not qty_col:
            # Look for text columns with numeric content (excluding serial number columns)
            for col in df.columns:
                col_lower = str(col).lower().strip()
                
                if (df[col].dtype == 'object' and col not in [desc_col, unit_col] and
                    not any(serial_pattern in col_lower for serial_pattern in serial_exclusion_patterns)):
                    numeric_count = 0
                    total_count = 0
                    for value in df[col].dropna().head(10):
                        total_count += 1
                        try:
                            float(str(value).replace(',', ''))
                            numeric_count += 1
                        except:
                            pass
                    
                    if total_count > 0 and numeric_count / total_count > 0.7:
                        qty_col = col
                        logger.info(f"✅ QUANTITY COLUMN DETECTED (text-numeric): '{col}'")
                        break
        
        if not qty_col:
            column_mappings["error_message"] = f"No quantity column found in sheet '{sheet_name}'"
            return column_mappings
        
        column_mappings["quantity_column"] = qty_col
        column_mappings["validation_status"] = "SUCCESS"
        
        logger.info(f"🎉 MANDATORY COLUMNS IDENTIFIED for '{sheet_name}':")
        logger.info(f"   📝 Description: '{desc_col}'")
        logger.info(f"   📏 Unit: '{unit_col}'") 
        logger.info(f"   🔢 Quantity: '{qty_col}'")
        
        return column_mappings

    def identify_mandatory_columns_from_headers(self, header_names: List[str], sheet_name: str) -> Dict:
        """Identify mandatory columns from actual header names (much faster and more accurate)"""
        
        column_mappings = {
            "description_column": None,
            "unit_column": None, 
            "quantity_column": None,
            "validation_status": "FAILED",
            "error_message": ""
        }
        
        # Serial number exclusion patterns - these columns should NEVER be identified as description/quantity columns
        serial_exclusion_patterns = ['item no', 'sl. no', 'sl.no', 'slno', 's.no', 's no', 'sr. no', 'sr.no', 'srno', 'serial', 'serial no', 'sno', 'item number', 'row number', 'line no', 'entry no']
        
        # Enhanced patterns with priority - exact matches first
        desc_patterns_priority = [
            ['description'],
            ['item description', 'work description', 'job description', 'activity description', 'particulars'],
            ['specification', 'scope of work', 'nature of work', 'work item', 'activity', 'task'],
            ['item', 'items', 'material', 'materials', 'work', 'details', 'desc', 'item name', 'work details', 'scope', 'type of work']
        ]
        
        unit_patterns_priority = [
            ['unit'],
            ['unit of measurement', 'u.o.m', 'uom', 'measuring unit'],
            ['units', 'measure', 'measurement', 'unit measure'],
            ['un', 'u/m', 'u.m', 'um']
        ]
        
        qty_patterns_priority = [
            ['quantity'],
            ['cumulative quantity', 'overall quantity', 'total quantity', 'net quantity', 'gross quantity', 'variation in quantity', 'variation in qty'],
            ['cumulative', 'overall qty', 'total qty', 'cum qty', 'cumulative qty', 'variation qty', 'var qty', 'variance qty'],
            # Removed problematic 'no', 'no.', 'nos', 'nos.' that match "Item No." - replaced with safer patterns
            ['qty', 'qnty', 'quant', 'volume', 'count', 'overall', 'cum', 'qtty', 'amount', 'variation', 'var', 'total']
        ]
        
        # Find description column
        desc_col_idx = None
        desc_col_name = None
        for priority_level, patterns in enumerate(desc_patterns_priority):
            for idx, header in enumerate(header_names):
                header_lower = str(header).lower().strip()
                
                # SKIP if this is clearly a serial number column
                if any(serial_pattern in header_lower for serial_pattern in serial_exclusion_patterns):
                    continue
                    
                if any(desc_name in header_lower for desc_name in patterns):
                    desc_col_idx = idx  # Store column index for extraction
                    desc_col_name = header  # Store header name for logging
                    logger.info(f"✅ DESCRIPTION COLUMN FOUND (Priority {priority_level+1}): '{header}' at index {idx}")
                    break
            if desc_col_idx is not None:
                break
        
        if desc_col_idx is None:
            column_mappings["error_message"] = f"No description column found in headers: {header_names}"
            return column_mappings
        
        column_mappings["description_column"] = desc_col_idx
        
        # Find unit column
        unit_col_idx = None
        unit_col_name = None
        for priority_level, patterns in enumerate(unit_patterns_priority):
            for idx, header in enumerate(header_names):
                header_lower = str(header).lower().strip()
                
                # SKIP if this is clearly a serial number column or already identified as description
                if (any(serial_pattern in header_lower for serial_pattern in serial_exclusion_patterns) or 
                    idx == desc_col_idx):
                    continue
                    
                if any(unit_name in header_lower for unit_name in patterns):
                    unit_col_idx = idx  # Store column index for extraction
                    unit_col_name = header  # Store header name for logging
                    logger.info(f"✅ UNIT COLUMN FOUND (Priority {priority_level+1}): '{header}' at index {idx}")
                    break
            if unit_col_idx is not None:
                break
        
        if unit_col_idx is None:
            column_mappings["error_message"] = f"No unit column found in headers: {header_names}"
            return column_mappings
        
        column_mappings["unit_column"] = unit_col_idx
        
        # Find quantity column
        qty_col_idx = None
        qty_col_name = None
        for priority_level, patterns in enumerate(qty_patterns_priority):
            for idx, header in enumerate(header_names):
                header_lower = str(header).lower().strip()
                
                # SKIP if this is clearly a serial number column or already identified columns
                if (any(serial_pattern in header_lower for serial_pattern in serial_exclusion_patterns) or 
                    idx == desc_col_idx or idx == unit_col_idx):
                    continue
                
                if any(qty_name in header_lower for qty_name in patterns):
                    qty_col_idx = idx  # Store column index for extraction
                    qty_col_name = header  # Store header name for logging
                    logger.info(f"✅ QUANTITY COLUMN FOUND (Priority {priority_level+1}): '{header}' at index {idx}")
                    break
            if qty_col_idx is not None:
                break
        
        if qty_col_idx is None:
            column_mappings["error_message"] = f"No quantity column found in headers: {header_names}"
            return column_mappings
        
        column_mappings["quantity_column"] = qty_col_idx
        column_mappings["validation_status"] = "SUCCESS"
        
        return column_mappings

    def analyze_worksheet_structure(self, ws, sheet_name: str, file_path: str) -> WorksheetAnalysis:
        """Comprehensive analysis of worksheet structure"""
        logger.info(f"📊 WORKSHEET ANALYSIS: Starting analysis of '{sheet_name}'")
        
        # Read data for analysis 
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            # Only handle merged cells if there are merged cells (optimization)
            if ws.merged_cells.ranges:
                logger.info(f"⚡ Handling {len(list(ws.merged_cells.ranges))} merged cell ranges")
                df = self.handle_merged_cells(ws, df)
        except Exception as e:
            logger.error(f"Error reading worksheet {sheet_name}: {str(e)}")
            df = pd.DataFrame()
        
        # Column mappings will be determined by AI during header selection
        
        # Find suspected headers
        suspected_headers = self.find_all_potential_headers(df, sheet_name)
        
        # DISABLED: AI header selection - using pure rule-based approach for speed testing
        # selected_header, ai_column_mappings = self.ai_select_best_header(suspected_headers, sheet_name)
        
        # Select best header using rule-based approach (highest confidence)
        selected_header = suspected_headers[0] if suspected_headers else None
        if selected_header:
            logger.info(f"🎯 RULE-BASED HEADER SELECTION: Using header at row {selected_header[0]+1} for '{sheet_name}'")
        
        # Use pure rule-based column detection with actual header names
        if not df.empty and selected_header:
            header_row, header_names, confidence = selected_header
            column_mappings = self.identify_mandatory_columns_from_headers(header_names, sheet_name)
        elif not df.empty:
            # Fallback to old method if no header selected
            column_mappings = self.identify_mandatory_columns(df, sheet_name)
        else:
            column_mappings = {"validation_status": "FAILED", "error_message": "Empty dataframe"}
        
        # SKIP BOQ region creation during preprocessing for speed
        boq_regions = []
        
        analysis = WorksheetAnalysis(
            name=sheet_name,
            dimensions=(0, 0),  # Not needed for speed
            frozen_panes=None,  # Not needed
            merged_cells=[],    # Not needed 
            suspected_headers=suspected_headers,
            boq_regions=boq_regions,  # Empty for speed
            data_density=0.0,   # Not needed
            has_wrapped_text=False,  # Not needed
            ai_analysis=None,
            column_mappings=column_mappings  # This is what we actually need
        )
        
        # AI analysis disabled for speed
        return analysis
    
    def is_likely_data_row(self, row_values: List[str]) -> bool:
        """Better check to distinguish data rows from header rows"""
        data_indicators = 0
        total_non_empty = 0
        
        for value in row_values:
            value_str = str(value).strip()
            if value_str and value_str.lower() != 'nan':
                total_non_empty += 1
                
                # Better data row indicators
                if re.match(r'^[A-Z]\s*\d+$', value_str):  # "A 2", "B 1" 
                    data_indicators += 2  # Strong indicator
                elif re.match(r'^\d+\.?\d*$', value_str):  # Pure numbers
                    data_indicators += 1
                elif re.match(r'^[A-Z]{2,}\d+$', value_str):  # "PCC003"
                    data_indicators += 2  # Strong indicator
                elif len(value_str) > 80:  # Very long descriptions
                    data_indicators += 1
                elif any(word in value_str.lower() for word in ['excavation', 'concrete', 'steel', 'masonry']):
                    data_indicators += 1  # Construction terms
        
        # Higher threshold for data row detection
        return (data_indicators / total_non_empty) > 0.6 if total_non_empty > 0 else False
    
    def find_all_potential_headers(self, df: pd.DataFrame, sheet_name: str) -> List[Tuple[int, List[str], float]]:
        """Better header detection with realistic confidence scores"""
        potential_headers = []
        
        # Search through first 25 rows for headers (minimal logging)
        search_range = min(25, len(df))
        
        for idx in range(search_range):
            row_data = df.iloc[idx].astype(str).fillna('')
            
            if row_data.str.strip().eq('').all():  # Skip completely empty rows
                continue
            
            # Get non-empty values for analysis
            non_empty_values = [cell for cell in row_data if cell and str(cell).strip() and str(cell).strip() != 'nan']
            
            if len(non_empty_values) < 2:  # Headers should have at least 2 columns
                continue
            
            # Better filter for data rows
            if self.is_likely_data_row(non_empty_values):
                continue
            
            # Better pattern matching
            score = 0
            matched_patterns = []
            logger.debug(f"🔍 Row {idx+1}: HEADER CANDIDATE - analyzing {len(non_empty_values)} cells for BOQ patterns")
            
            # Check each cell against BOQ patterns
            for cell_value in row_data:
                cell_str = str(cell_value).strip()
                if cell_str and cell_str != 'nan':
                    for pattern_name, patterns in self.boq_patterns.items():
                        for pattern in patterns:
                            if re.search(pattern, cell_str):
                                score += 1
                                matched_patterns.append(pattern_name)
                                logger.debug(f"🔍 Row {idx+1}: Found '{pattern_name}' pattern in '{cell_str[:20]}'")
                                break
            
            # More realistic confidence calculation
            non_empty_cells = len(non_empty_values)
            
            if non_empty_cells > 0 and score > 0:
                # Base confidence - never exceeds 0.8 without AI
                confidence = min(score / non_empty_cells, 0.8)
                
                # More conservative boosts
                unique_patterns = len(set(matched_patterns))
                if unique_patterns >= 4:
                    confidence *= 1.2
                elif unique_patterns >= 3:
                    confidence *= 1.1
                elif unique_patterns >= 2:
                    confidence *= 1.05
                
                # Boost for critical BOQ patterns
                critical_patterns = ['sl_no', 'description', 'unit', 'quantity', 'rate', 'amount']
                if any(pattern in matched_patterns for pattern in critical_patterns):
                    confidence *= 1.1
                
                # Cap at 0.9 (never 1.0 without AI validation)
                confidence = min(confidence, 0.9)
                
                # Special boost for headers with key BOQ indicators
                key_boq_indicators = ['sl_no', 'description', 'quantity', 'unit', 'rate', 'amount']
                key_matches = sum(1 for pattern in matched_patterns if pattern in key_boq_indicators)
                
                if key_matches >= 3:  # Has at least 3 key BOQ columns
                    confidence *= 1.5  # Significant boost
                    logger.info(f"🎯 STRONG BOQ HEADER detected with {key_matches} key indicators")
                
                # Lower threshold for acceptance
                if confidence > 0.20:  # Even lower threshold
                    headers = [str(cell).strip() for cell in row_data]
                    potential_headers.append((idx, headers, confidence))
                    logger.info(f"✅ HEADER CANDIDATE at row {idx+1}: confidence={confidence:.3f}, patterns={list(set(matched_patterns))}")
                    logger.info(f"   📋 Headers preview: {headers[:6]}...")
                    logger.info(f"   🔑 Key BOQ indicators: {key_matches}/6")
                else:
                    logger.debug(f"🔍 Row {idx+1}: Low confidence ({confidence:.3f}) - not adding as header candidate")
        
        # Smart deduplication: only one high-confidence header per region
        potential_headers = self.smart_deduplicate_headers(potential_headers)
        
        # Sort by confidence
        potential_headers.sort(key=lambda x: x[2], reverse=True)
        
        logger.info(f"🔍 HEADER DETECTION COMPLETE: Found {len(potential_headers)} potential header rows in '{sheet_name}'")
        if potential_headers:
            logger.info(f"   🥇 Top candidate: Row {potential_headers[0][0]+1} (confidence: {potential_headers[0][2]:.3f})")
            if len(potential_headers) > 1:
                logger.info(f"   🥈 Second candidate: Row {potential_headers[1][0]+1} (confidence: {potential_headers[1][2]:.3f})")
        else:
            logger.warning(f"⚠️ NO HEADERS FOUND in sheet '{sheet_name}' - this sheet may not contain BOQ data")
        
        return potential_headers
    
    def smart_deduplicate_headers(self, headers: List[Tuple[int, List[str], float]]) -> List[Tuple[int, List[str], float]]:
        """Remove conflicting headers - only keep one high-confidence header per region"""
        if not headers:
            return headers
        
        # Sort by confidence first
        headers.sort(key=lambda x: x[2], reverse=True)
        
        filtered = []
        used_rows = set()
        
        for row_idx, header_list, confidence in headers:
            # Check if this row conflicts with already selected headers
            conflicts = False
            for used_row in used_rows:
                if abs(row_idx - used_row) <= 5:  # Within 5 rows = same region
                    conflicts = True
                    break
            
            if not conflicts:
                filtered.append((row_idx, header_list, confidence))
                used_rows.add(row_idx)
        
        return filtered
    
    def call_gemini_api(self, prompt: str) -> Optional[Dict]:
        """Make a direct HTTP request to Gemini API"""
        if not self.api_key:
            return None
            
        # Prepare the API request
        url = f"{self.base_url}?key={self.api_key}"
        
        headers = {
            'Content-Type': 'application/json',
        }
        
        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": prompt
                        }
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 0.1,
                "topK": 1,
                "topP": 1,
                "maxOutputTokens": 1024,
            }
        }
        
        try:
            logger.info(f"Making Gemini API call #{self.api_call_count + 1}")
            response = requests.post(url, headers=headers, json=payload, timeout=30)
            
            # Update tracking variable
            self.api_call_count += 1
            
            if response.status_code == 200:
                result = response.json()
                if 'candidates' in result and len(result['candidates']) > 0:
                    text_response = result['candidates'][0]['content']['parts'][0]['text']
                    logger.info(f"Raw AI response: {text_response[:200]}...")  # Log first 200 chars
                    
                    # Clean up the response - remove markdown formatting
                    cleaned_response = text_response.strip()
                    if cleaned_response.startswith('```json'):
                        cleaned_response = cleaned_response[7:]
                    elif cleaned_response.startswith('```'):
                        cleaned_response = cleaned_response[3:]
                    if cleaned_response.endswith('```'):
                        cleaned_response = cleaned_response[:-3]
                    cleaned_response = cleaned_response.strip()
                    
                    # Remove any extra text before/after JSON
                    if '{' in cleaned_response and '}' in cleaned_response:
                        start = cleaned_response.find('{')
                        end = cleaned_response.rfind('}') + 1
                        cleaned_response = cleaned_response[start:end]
                    
                    try:
                        # Try to parse as JSON
                        ai_result = json.loads(cleaned_response)
                        logger.info(f"✅ AI JSON parsed successfully: {list(ai_result.keys())}")
                        return ai_result
                    except json.JSONDecodeError as e:
                        logger.warning(f"Failed to parse AI response as JSON: {str(e)}")
                        logger.warning(f"Response text: {cleaned_response[:500]}")
                        # Return a default structure if parsing fails
                        return {
                            "is_boq": True,  # Assume it's BOQ if we can't parse
                            "confidence": 0.5,
                            "boq_type": "construction",
                            "data_quality": "medium",
                            "extraction_recommendation": "extract",
                            "notes": "AI response parsing failed, using default values"
                        }
                else:
                    logger.error("No candidates in Gemini response")
                    return None
            else:
                logger.error(f"Gemini API error: {response.status_code} - {response.text}")
                return None
                
        except requests.exceptions.RequestException as e:
            logger.error(f"Request failed: {str(e)}")
            return None
        except Exception as e:
            logger.error(f"Unexpected error in Gemini API call: {str(e)}")
            return None
    
    def ai_select_best_header(self, suspected_headers: List[Tuple[int, List[str], float]], sheet_name: str) -> Tuple[Optional[Tuple[int, List[str], float]], Dict]:
        """Use AI to select the best header from top 2 candidates"""
        logger.info(f"🤖 AI HEADER SELECTION: Starting for sheet '{sheet_name}' with {len(suspected_headers)} candidates")
        
        if not self.ai_enabled:
            logger.warning(f"🤖 AI DISABLED - using top candidate if available")
            return (suspected_headers[0] if suspected_headers else None, {})
            
        if not suspected_headers:
            logger.warning(f"🤖 NO CANDIDATES provided for AI selection")
            return (None, {})
        
        # Take top 2 candidates for AI validation
        top_candidates = suspected_headers[:2]
        logger.info(f"🤖 AI analyzing top {len(top_candidates)} header candidates:")
        
        # Prepare candidates for AI
        candidates_info = []
        for i, (row_idx, headers, conf) in enumerate(top_candidates):
            candidate_info = {
                "option": i + 1,
                "row": row_idx + 1,
                "confidence": conf,
                "headers": headers[:6]  # First 6 headers only
            }
            candidates_info.append(candidate_info)
            logger.info(f"   Option {i+1}: Row {row_idx+1}, confidence={conf:.3f}, headers={headers[:4]}...")
        
        prompt = f"""You are analyzing BOQ (Bill of Quantities) headers. Select the best header and identify exact column names.

WORKSHEET: {sheet_name}
CANDIDATES:
{json.dumps(candidates_info, indent=2)}

TASK: Select the BEST BOQ header and identify these 3 mandatory columns:
1. DESCRIPTION column (contains work descriptions/items)  
2. UNIT column (contains units like cum, sqm, mt, kg, nos, etc.)
3. QUANTITY column (contains numeric quantities/volumes - may be named QTY, QUANTITY, OVERALL QTY, TOTAL QTY, etc.)

IMPORTANT: Column names must match EXACTLY as they appear in the candidate headers.

REQUIRED OUTPUT FORMAT:
```json
{{
  "selected_option": 1,
  "confidence": 0.95,
  "reasoning": "Option 1 has standard BOQ structure with clear columns",
  "column_mappings": {{
    "description_column": "EXACT_COLUMN_NAME_FROM_HEADERS",
    "unit_column": "EXACT_COLUMN_NAME_FROM_HEADERS",
    "quantity_column": "EXACT_COLUMN_NAME_FROM_HEADERS"
  }}
}}
```

CRITICAL: You MUST include the "column_mappings" field with exact column names from the headers.
If no valid BOQ header exists, set selected_option to 0.
Return ONLY the JSON, no extra text."""
        
        logger.info(f"🤖 Sending header selection request to AI...")
        ai_result = self.call_gemini_api(prompt)
        
        if ai_result:
            selected_option = ai_result.get('selected_option', 0)
            ai_confidence = ai_result.get('confidence', 0.5)
            ai_reasoning = ai_result.get('reasoning', 'No reasoning provided')
            ai_column_mappings = ai_result.get('column_mappings', {})
            
            logger.info(f"🤖 AI RESPONSE FULL: {ai_result}")
            logger.info(f"🤖 AI RESPONSE: selected_option={selected_option}, confidence={ai_confidence}")
            logger.info(f"🤖 AI REASONING: {ai_reasoning}")
            logger.info(f"🎯 AI COLUMN MAPPINGS: {ai_column_mappings}")
            
            if 1 <= selected_option <= len(top_candidates):
                selected_header = top_candidates[selected_option - 1]
                logger.info(f"✅ AI SELECTED HEADER: Option {selected_option} - Row {selected_header[0]+1} for sheet '{sheet_name}'")
                logger.info(f"   🎯 AI reasoning: {ai_reasoning}")
                
                # Initialize ai_column_mappings if empty
                if not ai_column_mappings:
                    ai_column_mappings = {}
                
                # Check if column_mappings is incomplete
                missing_cols = []
                for col in ["description_column", "unit_column", "quantity_column"]:
                    if not ai_column_mappings.get(col):
                        missing_cols.append(col)
                
                if missing_cols:
                    logger.warning(f"⚠️ Incomplete column_mappings in AI response, missing: {missing_cols}")
                    logger.warning(f"⚠️ Current AI mappings: {ai_column_mappings}")
                    logger.warning(f"⚠️ Completing from headers: {selected_header[1]}")
                    
                    complete_mappings = self.extract_columns_from_headers(selected_header[1], ai_reasoning)
                    logger.info(f"🔧 EXTRACTION RESULT: {complete_mappings}")
                    
                    # Merge AI results with extracted results (AI takes priority)
                    for col_type in ["description_column", "unit_column", "quantity_column"]:
                        if not ai_column_mappings.get(col_type):
                            ai_column_mappings[col_type] = complete_mappings.get(col_type)
                            logger.info(f"🔧 COMPLETED {col_type}: '{ai_column_mappings[col_type]}'")
                
                logger.info(f"🎯 FINAL AI COLUMN MAPPINGS: {ai_column_mappings}")
                
                # Validate column mappings
                validated_mappings = self.validate_ai_column_mappings(ai_column_mappings, selected_header[1])
                return (selected_header, validated_mappings)
            else:
                logger.warning(f"❌ AI REJECTED ALL HEADERS for sheet '{sheet_name}' - selected_option: {selected_option}")
                logger.warning(f"   🎯 AI reasoning: {ai_reasoning}")
                return (None, {})
        else:
            logger.warning(f"⚠️ AI header selection FAILED for sheet '{sheet_name}' - using top candidate as fallback")
            return (top_candidates[0], {})  # Fallback to highest confidence

    def extract_columns_from_headers(self, headers: List[str], ai_reasoning: str = "") -> Dict:
        """Smart extraction of column mappings from header list"""
        logger.info(f"🔍 EXTRACTING COLUMNS FROM HEADERS: {headers}")
        
        column_mappings = {}
        
        # Description column patterns (most common first)
        desc_patterns = ['description', 'item description', 'work description', 'particulars', 'activity', 'task', 'work item', 'specification', 'item']
        for header in headers:
            header_lower = str(header).lower()
            if any(pattern in header_lower for pattern in desc_patterns):
                column_mappings["description_column"] = header
                logger.info(f"📝 Found description column: '{header}'")
                break
        
        # Unit column patterns  
        unit_patterns = ['unit', 'uom', 'u.o.m', 'units', 'unit of measurement', 'measure']
        for header in headers:
            header_lower = str(header).lower()
            if any(pattern in header_lower for pattern in unit_patterns):
                column_mappings["unit_column"] = header
                logger.info(f"📏 Found unit column: '{header}'")
                break
        
        # Quantity column patterns (exact matches first, then partial)
        qty_exact_patterns = ['OVERALL QTY', 'TOTAL QTY', 'QTY', 'QUANTITY', 'CUMULATIVE', 'CUM QTY']
        qty_partial_patterns = ['qty', 'quantity', 'cumulative', 'overall qty', 'total qty', 'volume', 'count', 'number', 'overall', 'cum qty', 'cum']
        
        # First try exact matches (case sensitive)
        for header in headers:
            logger.info(f"🔍 Checking header '{header}' (type: {type(header)}) against exact patterns: {qty_exact_patterns}")
            if str(header) in qty_exact_patterns:
                column_mappings["quantity_column"] = header
                logger.info(f"🔢 Found quantity column (exact): '{header}'")
                break
            elif str(header).strip() == 'OVERALL QTY':  # Extra check with strip
                column_mappings["quantity_column"] = header
                logger.info(f"🔢 Found quantity column (exact with strip): '{header}'")
                break
        
        # If no exact match, try case-insensitive partial matches
        if "quantity_column" not in column_mappings:
            for header in headers:
                header_lower = str(header).lower()
                if any(pattern in header_lower for pattern in qty_partial_patterns):
                    column_mappings["quantity_column"] = header
                    logger.info(f"🔢 Found quantity column (partial): '{header}'")
                    break
        
        # If we still don't have all columns, try positional guessing based on typical BOQ structure
        if len(column_mappings) < 3:
            logger.warning(f"⚠️ Only found {len(column_mappings)}/3 columns, trying positional matching")
            
            # Typical BOQ order: S.NO, DESCRIPTION, UNIT, QTY, RATE, AMOUNT
            if len(headers) >= 3:
                if "description_column" not in column_mappings:
                    # Usually the longest header or second column
                    for i, header in enumerate(headers[1:4]):  # Check positions 1-3
                        if len(str(header)) > 8:  # Longer headers are usually descriptions
                            column_mappings["description_column"] = header
                            logger.info(f"📝 Guessed description column (position): '{header}'")
                            break
                
                if "unit_column" not in column_mappings:
                    # Look for short headers after description
                    for header in headers:
                        if len(str(header)) <= 6 and header not in column_mappings.values():
                            column_mappings["unit_column"] = header
                            logger.info(f"📏 Guessed unit column (short): '{header}'")
                            break
                
                if "quantity_column" not in column_mappings:
                    # Look for remaining numeric-sounding columns
                    for header in headers:
                        header_lower = str(header).lower()
                        if any(word in header_lower for word in ['qty', 'quant', 'total', 'cum', 'overall']) and header not in column_mappings.values():
                            column_mappings["quantity_column"] = header
                            logger.info(f"🔢 Guessed quantity column: '{header}'")
                            break
        
        logger.info(f"🎯 EXTRACTED COLUMNS: {column_mappings}")
        return column_mappings

    def validate_ai_column_mappings(self, ai_mappings: Dict, headers: List[str]) -> Dict:
        """Validate that AI-identified columns actually exist in the headers"""
        logger.info(f"🔍 VALIDATING AI COLUMN MAPPINGS: {ai_mappings}")
        logger.info(f"🔍 AVAILABLE HEADERS: {headers}")
        
        validated_mappings = {"validation_status": "FAILED", "error_message": ""}
        
        # Check that all required columns are provided by AI
        required_columns = ["description_column", "unit_column", "quantity_column"]
        missing_columns = []
        for col in required_columns:
            if col not in ai_mappings or not ai_mappings[col]:
                missing_columns.append(col)
        
        if missing_columns:
            validated_mappings["error_message"] = f"Missing columns: {missing_columns}. AI provided: {ai_mappings}"
            logger.error(f"❌ VALIDATION FAILED: {validated_mappings['error_message']}")
            return validated_mappings
        
        # Validate that AI-identified columns exist in actual headers
        for col_type, col_name in ai_mappings.items():
            if col_name not in headers:
                # Try case-insensitive match
                col_found = False
                for header in headers:
                    if str(header).lower() == str(col_name).lower():
                        ai_mappings[col_type] = header  # Use exact case from headers
                        col_found = True
                        break
                
                if not col_found:
                    validated_mappings["error_message"] = f"AI-identified column '{col_name}' not found in headers: {headers}"
                    return validated_mappings
        
        # Success
        validated_mappings = {
            "validation_status": "SUCCESS",
            "description_column": ai_mappings["description_column"],
            "unit_column": ai_mappings["unit_column"], 
            "quantity_column": ai_mappings["quantity_column"]
        }
        
        logger.info(f"✅ AI COLUMN MAPPINGS VALIDATED:")
        logger.info(f"   📝 Description: '{validated_mappings['description_column']}'")
        logger.info(f"   📏 Unit: '{validated_mappings['unit_column']}'")
        logger.info(f"   🔢 Quantity: '{validated_mappings['quantity_column']}'")
        
        return validated_mappings
    
    def ai_analyze_worksheet(self, df: pd.DataFrame, analysis: WorksheetAnalysis) -> Dict:
        """Simple validation of the final selected header"""
        # This is now just a simple confirmation, not used for validation
        return {"is_boq": True, "confidence": 0.8, "data_quality": "good"}
    
    def clean_headers(self, headers: List[str]) -> List[str]:
        """Clean and standardize header names, handling duplicates"""
        cleaned = []
        seen_headers = {}
        
        for i, header in enumerate(headers):
            if pd.isna(header):
                clean_header = f"Column_{i+1}"
            else:
                # Remove extra whitespace and newlines, limit length
                clean_header = re.sub(r'\s+', ' ', str(header).strip())
                # Limit header length to 50 characters for readability
                if len(clean_header) > 50:
                    clean_header = clean_header[:47] + "..."
            
            # Handle duplicates by adding a number suffix
            original_header = clean_header
            counter = 1
            while clean_header in seen_headers:
                clean_header = f"{original_header}_{counter}"
                counter += 1
            
            seen_headers[clean_header] = True
            cleaned.append(clean_header)
        
        return cleaned
    
    def extract_table_boundaries(self, df: pd.DataFrame, header_row: int) -> Tuple[int, int]:
        """Extract table start and end rows, continuing until the true end of data"""
        start_row = header_row + 1
        end_row = -1  # Default to -1 if no data is found below the header
        
        logger.info(f"📊 BOUNDARY DETECTION: Header at row {header_row}, checking data from row {start_row} to {len(df)-1}")

        # Scan from the bottom up to find the last row with any data
        for idx in range(len(df) - 1, start_row - 1, -1):
            row_data = df.iloc[idx]
            # More flexible check - consider row not empty if it has any non-null, non-empty values
            has_data = False
            for cell in row_data:
                cell_str = str(cell).strip()
                if cell_str and cell_str.lower() not in ['nan', 'none', '']:
                    has_data = True
                    break
            
            if has_data:
                end_row = idx
                logger.info(f"📊 BOUNDARY DETECTION: Found last data row at {idx}")
                break
        
        if end_row == -1:
            # If no data found, check if there are at least a few rows below header
            max_check = min(start_row + 10, len(df) - 1)  # Check up to 10 rows below header
            logger.warning(f"📊 BOUNDARY DETECTION: No clear data end found, checking first {max_check - start_row + 1} rows after header")
            
            # Look for any row with some content
            for idx in range(start_row, max_check + 1):
                if idx < len(df):
                    row_data = df.iloc[idx]
                    non_empty_cells = sum(1 for cell in row_data if str(cell).strip() and str(cell).strip().lower() not in ['nan', 'none', ''])
                    if non_empty_cells >= 2:  # At least 2 non-empty cells
                        end_row = max(idx + 5, max_check)  # Give some buffer
                        logger.info(f"📊 BOUNDARY DETECTION: Using fallback end row {end_row} based on content detection")
                        break
                        
        return start_row, end_row

    def extract_boq_table(self, file_path: str, ws_analysis: WorksheetAnalysis, region: Dict) -> Optional[BOQTable]:
        """Extract a BOQ table from a specific region"""
        try:
            # Read worksheet data
            df = pd.read_excel(file_path, sheet_name=ws_analysis.name, header=None, engine='openpyxl')
            
            # Handle merged cells
            wb = openpyxl.load_workbook(file_path, data_only=False)
            ws = wb[ws_analysis.name]
            df = self.handle_merged_cells(ws, df)
            
            # Extract the specific region
            header_row = region['header_row']
            start_row = region['start_row']
            end_row = region['end_row']
            
            headers = region['headers']
            data_df = df.iloc[start_row:end_row+1].copy()
            
            # Clean up data
            clean_headers = self.clean_headers(headers)
            data_df.columns = clean_headers[:len(data_df.columns)]
            data_df = data_df.dropna(how='all')  # Remove completely empty rows
            
            # Collect cell details for verification
            cell_details = []
            for row_idx in range(header_row, min(header_row + 10, end_row + 1)):
                for col_idx in range(min(10, len(clean_headers))):
                    try:
                        cell_coord = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                        cell_info = self.analyze_cell_details(ws, cell_coord)
                        cell_details.append(cell_info)
                    except:
                        continue
            
            # AI validation is now header-level, not worksheet-level
            ai_validated = region.get('ai_selected', False)
            
            extraction_notes = [
                f"Header selected by AI from {len(ws_analysis.suspected_headers)} candidates" if ai_validated else f"Header confidence: {region['confidence']:.3f}",
                f"Contains {len(data_df)} data rows",
                f"AI header selection: {'YES' if ai_validated else 'NO'}"
            ]
            
            boq_table = BOQTable(
                worksheet_name=ws_analysis.name,
                start_row=header_row,
                start_col=0,
                end_row=end_row,
                end_col=len(clean_headers)-1,
                headers=clean_headers,
                data=data_df,
                confidence_score=region['confidence'],
                ai_validated=ai_validated,  # Now means "AI selected this header"
                extraction_notes=extraction_notes,
                cell_details=cell_details
            )
            
            logger.info(f"Successfully extracted BOQ table from {ws_analysis.name}: {len(data_df)} rows, AI selected: {ai_validated}")
            return boq_table
            
        except Exception as e:
            logger.error(f"Error extracting BOQ table from {ws_analysis.name}: {str(e)}")
            return None
    
    def process_excel_file(self, file_path: str) -> FileAnalysis:
        """Process a complete Excel file and return analysis results"""
        start_time = time.time()
        logger.info(f"🚀 STARTING FILE PROCESSING: {os.path.basename(file_path)}")
        
        # Validate file exists and is readable
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Get file info
        file_size = os.path.getsize(file_path)
        logger.info(f"📄 File size: {file_size:,} bytes ({file_size/(1024*1024):.1f} MB)")
        
        # Load workbook with error handling
        try:
            logger.info(f"📖 Loading Excel workbook...")
            wb = openpyxl.load_workbook(file_path, data_only=False)
            logger.info(f"✅ Workbook loaded successfully")
        except Exception as e:
            logger.error(f"❌ Failed to load workbook: {str(e)}")
            raise Exception(f"Cannot open Excel file. File may be corrupted or password protected: {str(e)}")
        
        # ULTRA-FAST: Load ALL sheets in single operation (massive speed improvement)
        logger.info(f"📋 Found {len(wb.sheetnames)} worksheets: {wb.sheetnames}")
        logger.info(f"🚀 USING ULTRA-FAST BULK PROCESSING: Single file read for all sheets")
        
        try:
            # Use the new ultra-fast method (single file read)
            worksheet_analyses = self.analyze_all_worksheets_ultra_fast(file_path)
            
            # Extract BOQ tables from successful analyses
            boq_tables = []
            for ws_analysis in worksheet_analyses:
                if ws_analysis.boq_regions:
                    region = ws_analysis.boq_regions[0]  # Only 1 region per worksheet
                    if region['data_rows'] > 0:
                        logger.info(f"📊 Extracting BOQ table from '{ws_analysis.name}' with {region['data_rows']} data rows")
                        boq_table = self.extract_boq_table(file_path, ws_analysis, region)
                        if boq_table:
                            boq_tables.append(boq_table)
                            logger.info(f"✅ Successfully created BOQ table for '{ws_analysis.name}' ({len(boq_table.data)} rows)")
                        else:
                            logger.warning(f"⚠️ Failed to extract BOQ table from '{ws_analysis.name}'")
                    else:
                        logger.warning(f"⚠️ Skipping BOQ region in '{ws_analysis.name}' - contains no data rows")
                else:
                    logger.info(f"📊 No BOQ regions found in worksheet '{ws_analysis.name}'")
                    
        except Exception as e:
            logger.error(f"❌ BATCH ANALYSIS FAILED: {str(e)}")
            logger.warning(f"⚠️ FALLING BACK to individual worksheet analysis...")
            
            # Fallback to original individual analysis
            worksheet_analyses = []
            boq_tables = []
            
            for i, sheet_name in enumerate(wb.sheetnames, 1):
                try:
                    logger.info(f"📋 PROCESSING WORKSHEET {i}/{len(wb.sheetnames)}: '{sheet_name}' (FALLBACK)")
                    ws = wb[sheet_name]
                    ws_analysis = self.analyze_worksheet_structure(ws, sheet_name, file_path)
                    worksheet_analyses.append(ws_analysis)
                    
                    # Each worksheet now has max 1 BOQ region (AI-selected header)
                    if ws_analysis.boq_regions:
                        region = ws_analysis.boq_regions[0]  # Only 1 region per worksheet
                        if region['data_rows'] > 0:
                            logger.info(f"📊 Extracting BOQ table from region with {region['data_rows']} data rows")
                            boq_table = self.extract_boq_table(file_path, ws_analysis, region)
                            if boq_table:
                                boq_tables.append(boq_table)
                                logger.info(f"✅ Successfully created BOQ table for '{sheet_name}' ({len(boq_table.data)} rows)")
                            else:
                                logger.warning(f"⚠️ Failed to extract BOQ table from '{sheet_name}'")
                        else:
                            logger.warning(f"⚠️ Skipping BOQ region in '{sheet_name}' - contains no data rows")
                    else:
                        logger.info(f"📊 No BOQ regions found in worksheet '{sheet_name}'")
                        
                except Exception as e:
                    logger.error(f"❌ Error processing worksheet '{sheet_name}': {str(e)}")
                    continue
        
        processing_time = time.time() - start_time
        
        # Create summary
        worksheets_with_boq = len([ws for ws in worksheet_analyses if ws.boq_regions])
        high_confidence_tables = len([bt for bt in boq_tables if bt.confidence_score > 0.7])
        ai_validated_tables = len([bt for bt in boq_tables if bt.ai_validated]) if self.ai_enabled else 0
        
        extraction_summary = {
            'total_worksheets': len(worksheet_analyses),
            'worksheets_with_boq': worksheets_with_boq,
            'total_boq_tables': len(boq_tables),
            'high_confidence_tables': high_confidence_tables,
            'ai_validated_tables': ai_validated_tables,
            'processing_time_seconds': processing_time
        }
        
        logger.info(f"🎉 FILE PROCESSING COMPLETE!")
        logger.info(f"   📊 Total worksheets: {len(worksheet_analyses)}")
        logger.info(f"   ✅ Worksheets with BOQ: {worksheets_with_boq}")
        logger.info(f"   📋 BOQ tables created: {len(boq_tables)}")
        logger.info(f"   🎯 High confidence tables: {high_confidence_tables}")
        logger.info(f"   🤖 AI validated tables: {ai_validated_tables}")
        logger.info(f"   ⏱️ Processing time: {processing_time:.2f} seconds")
        
        file_analysis = FileAnalysis(
            filename=os.path.basename(file_path),
            file_size=file_size,
            worksheets=worksheet_analyses,
            total_boq_tables=len(boq_tables),
            extraction_summary=extraction_summary,
            processing_time=processing_time
        )
        
        return file_analysis

    def analyze_all_worksheets_batch(self, wb, file_path: str) -> List[WorksheetAnalysis]:
        """
        OPTIMIZED: Analyze ALL worksheets with a single batch AI call for header selection
        
        This method:
        1. Collects suspected headers from all worksheets
        2. Makes ONE AI call to select best headers for ALL sheets
        3. Processes column mappings for all sheets
        
        Args:
            wb: Openpyxl workbook object
            file_path: Path to the Excel file
            
        Returns:
            List of WorksheetAnalysis objects for all sheets
        """
        logger.info(f"🚀 BATCH WORKSHEET ANALYSIS: Starting analysis of ALL {len(wb.sheetnames)} worksheets")
        
        # Step 1: BATCH DATA COLLECTION from all worksheets simultaneously
        all_worksheet_data = []
        logger.info(f"📊 Batch processing {len(wb.sheetnames)} worksheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            
            try:
                ws = wb[sheet_name]
                
                # Read data for analysis
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
                    # Only handle merged cells if there are merged cells (optimization)
                    if ws.merged_cells.ranges:
                        df = self.handle_merged_cells(ws, df)
                except Exception as e:
                    logger.error(f"Error reading worksheet {sheet_name}: {str(e)}")
                    df = pd.DataFrame()
                
                # Find suspected headers for this worksheet
                suspected_headers = self.find_all_potential_headers(df, sheet_name)
                
                # Store MINIMAL data for batch processing (only what's needed)
                worksheet_data = {
                    'sheet_name': sheet_name,
                    'df': df,
                    'suspected_headers': suspected_headers
                }
                all_worksheet_data.append(worksheet_data)
                
                
            except Exception as e:
                logger.error(f"❌ Error collecting data for worksheet '{sheet_name}': {e}")
                continue
        
        # DISABLED: Batch AI call - using pure rule-based approach for speed testing
        # logger.info(f"🤖 BATCH AI HEADER SELECTION: Processing {len(all_worksheet_data)} worksheets in single AI call")
        # batch_ai_results = self.ai_select_best_headers_batch(all_worksheet_data)
        
        logger.info(f"🎯 RULE-BASED BATCH PROCESSING: Processing {len(all_worksheet_data)} worksheets (AI disabled for speed testing)")
        batch_ai_results = {}
        
        # Step 3: BATCH FINALIZATION - Build all WorksheetAnalysis objects
        worksheet_analyses = []
        logger.info(f"🔧 Finalizing analysis for all {len(all_worksheet_data)} worksheets")
        
        for ws_data in all_worksheet_data:
            sheet_name = ws_data['sheet_name']
            
            # RULE-BASED PROCESSING (AI disabled for speed testing)
            suspected_headers = ws_data['suspected_headers']
            selected_header = suspected_headers[0] if suspected_headers else None
            
            # Use header-based column detection for accuracy
            if not ws_data['df'].empty and selected_header:
                header_row, header_names, confidence = selected_header
                column_mappings = self.identify_mandatory_columns_from_headers(header_names, sheet_name)
            elif not ws_data['df'].empty:
                # Fallback to old method if no header selected
                column_mappings = self.identify_mandatory_columns(ws_data['df'], sheet_name)
            else:
                column_mappings = {"validation_status": "FAILED", "error_message": "Empty dataframe"}
            
            # SKIP BOQ region creation during preprocessing for speed
            # BOQ regions will be created later during extraction phase
            boq_regions = []
            ai_analysis = None
            
            # Create MINIMAL WorksheetAnalysis object (only essential data for speed)
            ws_analysis = WorksheetAnalysis(
                name=sheet_name,
                dimensions=(0, 0),  # Not needed for header identification
                frozen_panes=None,  # Not needed 
                merged_cells=[],    # Not needed
                suspected_headers=ws_data['suspected_headers'],
                boq_regions=boq_regions,  # Empty for speed
                data_density=0.0,   # Not needed
                has_wrapped_text=False,  # Not needed
                ai_analysis=ai_analysis,  # None for speed
                column_mappings=column_mappings  # This is the key info we need
            )
            
            worksheet_analyses.append(ws_analysis)
        
        logger.info(f"🎉 FAST PREPROCESSING COMPLETE: {len(worksheet_analyses)} worksheets processed")
        return worksheet_analyses

    def analyze_all_worksheets_ultra_fast(self, file_path: str) -> List[WorksheetAnalysis]:
        """
        ULTRA-FAST BULK PROCESSING: Load ALL sheets at once and process them together
        This eliminates per-sheet file I/O overhead
        """
        start_time = time.time()
        logger.info(f"🚀 ULTRA-FAST BULK PROCESSING: Starting analysis of file '{file_path}'")
        
        try:
            # SINGLE FILE READ - Load ALL sheets at once (massive speed improvement)
            logger.info(f"📊 Loading ALL sheets in single operation...")
            all_sheets_dict = pd.read_excel(file_path, sheet_name=None, header=None, engine='openpyxl')
            sheet_names = list(all_sheets_dict.keys())
            logger.info(f"📊 Loaded {len(sheet_names)} sheets in {time.time() - start_time:.2f}s: {sheet_names}")
            
            # BULK PROCESSING - Process all sheets together
            worksheet_analyses = []
            
            for sheet_name, df in all_sheets_dict.items():
                # Skip empty sheets
                if df.empty:
                    continue
                
                # Find headers and detect columns (minimal logging)
                suspected_headers = self.find_all_potential_headers(df, sheet_name)
                selected_header = suspected_headers[0] if suspected_headers else None
                
                if selected_header:
                    header_row, header_names, confidence = selected_header
                    column_mappings = self.identify_mandatory_columns_from_headers(header_names, sheet_name)
                else:
                    column_mappings = {"validation_status": "FAILED", "error_message": "No headers found"}
                
                # Create minimal analysis object
                ws_analysis = WorksheetAnalysis(
                    name=sheet_name,
                    dimensions=(0, 0),
                    frozen_panes=None,
                    merged_cells=[],
                    suspected_headers=suspected_headers,
                    boq_regions=[],  # Empty for speed
                    data_density=0.0,
                    has_wrapped_text=False,
                    ai_analysis=None,
                    column_mappings=column_mappings
                )
                
                worksheet_analyses.append(ws_analysis)
            
            total_time = time.time() - start_time
            logger.info(f"🎉 ULTRA-FAST PROCESSING COMPLETE: {len(worksheet_analyses)} sheets in {total_time:.2f}s ({total_time/len(worksheet_analyses):.3f}s per sheet)")
            return worksheet_analyses
            
        except Exception as e:
            logger.error(f"❌ ULTRA-FAST PROCESSING ERROR: {str(e)}")
            # Fallback to regular processing
            logger.info(f"🔄 Falling back to regular batch processing...")
            return self.analyze_all_worksheets_batch(openpyxl.load_workbook(file_path, read_only=True), file_path)

    def ai_select_best_headers_batch(self, all_worksheet_data: List[Dict]) -> Dict:
        """
        Make a single AI call to select best headers for ALL worksheets
        
        Args:
            all_worksheet_data: List of worksheet data dictionaries
            
        Returns:
            Dictionary mapping sheet names to their AI results
        """
        if not self.ai_enabled:
            logger.warning(f"🤖 AI DISABLED - using top candidates for all sheets")
            results = {}
            for ws_data in all_worksheet_data:
                sheet_name = ws_data['sheet_name']
                suspected_headers = ws_data['suspected_headers']
                if suspected_headers:
                    selected_header = suspected_headers[0]  # Use top candidate
                    results[sheet_name] = {
                        'selected_header': selected_header,
                        'column_mappings': self.identify_mandatory_columns(ws_data['df'], sheet_name)
                    }
            return results
        
        # Prepare batch prompt with all worksheets
        batch_prompt = self._create_batch_header_selection_prompt(all_worksheet_data)
        
        logger.info(f"🤖 Making SINGLE AI call for {len(all_worksheet_data)} worksheets")
        
        try:
            # Use the existing call_gemini_api method, but handle the response carefully
            response = self.call_gemini_api(batch_prompt)
            logger.info(f"🔍 DEBUG: Raw AI response: {response}")
            
            if response and isinstance(response, dict):
                # Check if it's a valid batch response or a default fallback response
                if 'is_boq' in response and 'confidence' in response:
                    # This is the default response structure from call_gemini_api when JSON parsing fails
                    # Fall back to rule-based selection
                    logger.warning("🤖 AI returned default response structure, falling back to rule-based selection")
                    return self._fallback_header_selection(all_worksheet_data)
                else:
                    # This looks like a valid batch response
                    logger.info(f"🔍 DEBUG: Parsing valid batch response with keys: {list(response.keys())}")
                    return self._parse_batch_ai_response_dict(response, all_worksheet_data)
            else:
                logger.warning("🤖 AI request failed, falling back to rule-based selection for all sheets")
                return self._fallback_header_selection(all_worksheet_data)
                
        except Exception as e:
            logger.error(f"🤖 AI batch request failed: {e}")
            return self._fallback_header_selection(all_worksheet_data)

    def _create_batch_header_selection_prompt(self, all_worksheet_data: List[Dict]) -> str:
        """Create a prompt for batch header selection across all worksheets"""
        prompt_parts = [
            "You are analyzing multiple Excel worksheets from a BOQ (Bill of Quantities) file.",
            "For each worksheet, select the BEST header row that contains column names like 'Description', 'Unit', 'Quantity', etc.",
            "",
            "WORKSHEETS TO ANALYZE:",
            ""
        ]
        
        for i, ws_data in enumerate(all_worksheet_data, 1):
            sheet_name = ws_data['sheet_name']
            suspected_headers = ws_data['suspected_headers']
            
            prompt_parts.append(f"=== WORKSHEET {i}: '{sheet_name}' ===")
            prompt_parts.append(f"Suspected headers found: {len(suspected_headers)}")
            
            for j, (row_num, headers, confidence) in enumerate(suspected_headers[:3], 1):  # Top 3 candidates
                headers_str = " | ".join(str(h) for h in headers[:10])  # Limit display
                prompt_parts.append(f"  Candidate {j} (Row {row_num}, Score: {confidence:.2f}): {headers_str}")
            
            prompt_parts.append("")
        
        prompt_parts.extend([
            "INSTRUCTIONS:",
            "1. For each worksheet, select the BEST header candidate that contains BOQ columns",
            "2. Identify exact column names for Description, Unit, and Quantity from the selected header",
            "3. MUST set validation_status to SUCCESS if valid columns are found",
            "4. Use the exact worksheet names provided above",
            "5. Respond in JSON format with this EXACT structure:",
            "",
            "{",
            '  "EXACT_WORKSHEET_NAME": {',
            '    "selected_candidate": 1,',
            '    "confidence": 0.95,',
            '    "reasoning": "Clear BOQ headers found",',
            '    "column_mappings": {',
            '      "description_column": "EXACT_COLUMN_NAME_FROM_HEADERS",',
            '      "unit_column": "EXACT_COLUMN_NAME_FROM_HEADERS",',
            '      "quantity_column": "EXACT_COLUMN_NAME_FROM_HEADERS",',
            '      "validation_status": "SUCCESS"',
            '    }',
            '  }',
            '}',
            "",
            "CRITICAL REQUIREMENTS:",
            "- Use EXACT worksheet names shown above",
            "- Use EXACT column names from the header candidates", 
            "- ALWAYS include validation_status field",
            "- Set validation_status to SUCCESS if valid BOQ columns found",
            "- If no valid BOQ headers exist, set selected_candidate to 0 and validation_status to FAILED"
        ])
        
        return "\n".join(prompt_parts)

    def _parse_batch_ai_response_dict(self, ai_response: Dict, all_worksheet_data: List[Dict]) -> Dict:
        """Parse the batch AI response when it's already a dictionary"""
        results = {}
        
        try:
            # The response is already parsed as a dictionary
            parsed_response = ai_response
            
            for ws_data in all_worksheet_data:
                sheet_name = ws_data['sheet_name']
                
                if sheet_name in parsed_response:
                    sheet_ai_result = parsed_response[sheet_name]
                    selected_candidate_num = sheet_ai_result.get('selected_candidate', 1)
                    
                    # Get the selected header
                    suspected_headers = ws_data['suspected_headers']
                    if suspected_headers and 1 <= selected_candidate_num <= len(suspected_headers):
                        selected_header = suspected_headers[selected_candidate_num - 1]
                        
                        results[sheet_name] = {
                            'selected_header': selected_header,
                            'column_mappings': sheet_ai_result.get('column_mappings', {"validation_status": "FAILED"}),
                            'reasoning': sheet_ai_result.get('reasoning', 'AI selected'),
                            'confidence': sheet_ai_result.get('confidence', 0.8)
                        }
                    else:
                        logger.warning(f"Invalid candidate number for sheet '{sheet_name}', using fallback")
                        results[sheet_name] = self._fallback_single_sheet(ws_data)
                else:
                    logger.warning(f"No AI result for sheet '{sheet_name}', using fallback")
                    results[sheet_name] = self._fallback_single_sheet(ws_data)
                    
        except Exception as e:
            logger.error(f"Failed to parse AI dict response: {e}")
            return self._fallback_header_selection(all_worksheet_data)
        
        return results

    def _parse_batch_ai_response(self, ai_response: str, all_worksheet_data: List[Dict]) -> Dict:
        """Parse the batch AI response and return results for each worksheet"""
        results = {}
        
        try:
            # Try to parse JSON response
            parsed_response = json.loads(ai_response.strip())
            
            for ws_data in all_worksheet_data:
                sheet_name = ws_data['sheet_name']
                
                if sheet_name in parsed_response:
                    sheet_ai_result = parsed_response[sheet_name]
                    selected_candidate_num = sheet_ai_result.get('selected_candidate', 1)
                    
                    # Get the selected header
                    suspected_headers = ws_data['suspected_headers']
                    if suspected_headers and 1 <= selected_candidate_num <= len(suspected_headers):
                        selected_header = suspected_headers[selected_candidate_num - 1]
                        
                        results[sheet_name] = {
                            'selected_header': selected_header,
                            'column_mappings': sheet_ai_result.get('column_mappings', {"validation_status": "FAILED"}),
                            'reasoning': sheet_ai_result.get('reasoning', 'AI selected'),
                            'confidence': sheet_ai_result.get('confidence', 0.8)
                        }
                    else:
                        logger.warning(f"Invalid candidate number for sheet '{sheet_name}', using fallback")
                        results[sheet_name] = self._fallback_single_sheet(ws_data)
                else:
                    logger.warning(f"No AI result for sheet '{sheet_name}', using fallback")
                    results[sheet_name] = self._fallback_single_sheet(ws_data)
                    
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse AI JSON response: {e}")
            return self._fallback_header_selection(all_worksheet_data)
        
        return results

    def _fallback_header_selection(self, all_worksheet_data: List[Dict]) -> Dict:
        """Fallback to rule-based header selection for all worksheets"""
        results = {}
        for ws_data in all_worksheet_data:
            results[ws_data['sheet_name']] = self._fallback_single_sheet(ws_data)
        return results

    def _fallback_single_sheet(self, ws_data: Dict) -> Dict:
        """Fallback selection for a single worksheet"""
        sheet_name = ws_data['sheet_name']
        suspected_headers = ws_data['suspected_headers']
        
        selected_header = suspected_headers[0] if suspected_headers else None
        column_mappings = self.identify_mandatory_columns(ws_data['df'], sheet_name) if not ws_data['df'].empty else {"validation_status": "FAILED"}
        
        return {
            'selected_header': selected_header,
            'column_mappings': column_mappings,
            'reasoning': 'Fallback to top candidate',
            'confidence': 0.6
        }


def run_initial_analysis(file_path: str, gemini_api_key: str = None) -> FileAnalysis:
    """
    Instantiates and runs the BOQ pre-processor on a given file using OPTIMIZED BATCH ANALYSIS.
    
    This function now uses batch processing to analyze ALL worksheets with a single AI call
    instead of processing them one by one. This significantly reduces processing time and API costs.
    
    Args:
        file_path: The path to the uploaded Excel file.
        gemini_api_key: The API key for Gemini services.
        
    Returns:
        A FileAnalysis object containing the results.
    """
    # FIXED: Pass the API key properly to the extractor
    extractor = AdvancedBOQExtractor(gemini_api_key=gemini_api_key)
    analysis_results = extractor.process_excel_file(file_path)
    return analysis_results

def run_initial_analysis_legacy(file_path: str, gemini_api_key: str = None) -> FileAnalysis:
    """
    LEGACY: Individual worksheet analysis (one AI call per sheet).
    
    This is the old method that processes worksheets one by one.
    Use run_initial_analysis() for the optimized batch version.
    """
    extractor = AdvancedBOQExtractor(gemini_api_key=gemini_api_key)
    # Force the extractor to use individual analysis by temporarily disabling batch
    extractor._use_batch_analysis = False
    analysis_results = extractor.process_excel_file(file_path)
    return analysis_results