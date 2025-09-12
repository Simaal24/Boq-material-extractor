"""
BOQ Orchestrator Module

This module contains the main pipeline logic that coordinates all stages of BOQ processing.
It runs Stage 1 (detection), Stage 2 (verification), and Stage 3 (extraction) in sequence,
then provides functions for final summarization.
"""

import pandas as pd
import asyncio
import logging
from typing import List, Dict
from io import BytesIO

# Import all the engine parts we have built
from .stage1_detector import Stage1EnhancedHeaderDetector
from .stage2_verifier import Stage2EnhancedHeaderVerifier
from .stage3_extractor import Stage3EnhancedContextInheritance
from .summarizer import generate_summary_dataframe
from .data_structures import ExtractedLineItem, VerifiedHeader

logger = logging.getLogger(__name__)

async def run_extraction_pipeline(file_path: str, selected_sheets: List[str], api_key: str, cached_analysis=None) -> Dict:
    """
    Orchestrates the entire 3-stage BOQ extraction process for selected sheets.
    
    Args:
        file_path: Path to the Excel file
        selected_sheets: List of worksheet names to process
        api_key: Gemini API key (passed to ensure client is initialized)
        cached_analysis: Pre-computed analysis result (optional, avoids re-running analysis)
        
    Returns:
        Dictionary with results for each sheet
    """
    logger.info(f"🚀 ORCHESTRATOR: Starting 3-stage extraction for {len(selected_sheets)} sheets from {file_path}")
    
    # Import preprocessor for column identification
    from .pre_processor import AdvancedBOQExtractor
    
    # Initialize all processing stages
    stage1 = Stage1EnhancedHeaderDetector()
    stage2 = Stage2EnhancedHeaderVerifier()  # Uses centralized Gemini client
    stage3 = Stage3EnhancedContextInheritance()  # Uses centralized Gemini client
    
    all_results = {}

    try:
        xls = pd.ExcelFile(file_path)
        
        # Initialize preprocessor for column detection
        preprocessor = AdvancedBOQExtractor(gemini_api_key=api_key)
        
        for sheet_name in selected_sheets:
            logger.info(f"🔄 ORCHESTRATOR: Processing sheet: {sheet_name}")
            
            if sheet_name not in xls.sheet_names:
                logger.warning(f"❌ Sheet {sheet_name} not found in file")
                all_results[sheet_name] = {"error": f"Sheet {sheet_name} not found in file"}
                continue
                
            # OPTIMIZATION: Use cached analysis instead of re-running analysis
            logger.info(f"🔍 RETRIEVING COLUMN MAPPINGS: Using cached analysis for '{sheet_name}'")
            
            # Get analysis data from cache for this specific sheet
            analysis = None
            if cached_analysis:
                # Find the worksheet analysis in the cached results
                for ws_analysis in cached_analysis.worksheets:
                    if ws_analysis.name == sheet_name:
                        analysis = ws_analysis
                        logger.info(f"✅ Found CACHED analysis for sheet '{sheet_name}'")
                        break
            
            if not analysis:
                logger.warning(f"⚠️ No cached analysis found for '{sheet_name}', falling back to fresh analysis")
                # Fallback: run individual analysis for this sheet only
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, data_only=True)
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        analysis = preprocessor.analyze_worksheet_structure(ws, sheet_name, file_path)
                except Exception as e:
                    logger.error(f"Error loading worksheet for fallback analysis: {e}")
            
            # Read DataFrame using the correct header row from cached analysis
            header_row = 0  # Default
            if analysis and analysis.boq_regions:
                header_row = analysis.boq_regions[0]['header_row']
                logger.info(f"📊 Using CACHED header row: {header_row}")
            
            try:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
                logger.info(f"📊 DataFrame columns after reading with cached header row {header_row}: {list(df.columns)}")
            except Exception as e:
                logger.warning(f"⚠️ Failed to read with header row {header_row}, using default: {e}")
                df = pd.read_excel(xls, sheet_name=sheet_name)
            
            if df.empty:
                logger.warning(f"❌ Sheet {sheet_name} is empty")
                all_results[sheet_name] = {"error": f"Sheet {sheet_name} is empty"}
                continue

            # Get column mappings from cached analysis
            if analysis and hasattr(analysis, 'column_mappings'):
                column_mappings = analysis.column_mappings
                logger.info(f"✅ Using CACHED column mappings for '{sheet_name}'")
            else:
                # Fallback to manual detection on the properly read DataFrame
                logger.warning(f"⚠️ No cached column mappings, falling back to manual detection for '{sheet_name}'")
                column_mappings = preprocessor.identify_mandatory_columns(df, sheet_name)
            
            if column_mappings.get("validation_status") != "SUCCESS":
                error_msg = column_mappings.get("error_message", "Failed to identify mandatory columns")
                logger.error(f"❌ COLUMN MAPPINGS FAILED for '{sheet_name}': {error_msg}")
                all_results[sheet_name] = {"error": f"Missing required BOQ columns: {error_msg}"}
                continue
            
            # Extract column names from mappings
            desc_col = column_mappings["description_column"]
            unit_col = column_mappings["unit_column"] 
            quantity_col = column_mappings["quantity_column"]
            
            logger.warning(f"✅ AI-IDENTIFIED COLUMNS FOUND for '{sheet_name}':")
            logger.warning(f"   📝 Description: '{desc_col}'")
            logger.warning(f"   📏 Unit: '{unit_col}'")
            logger.warning(f"   🔢 Quantity: '{quantity_col}'")

            # STAGE 1: Rule-based header detection and chunking
            logger.warning(f"🚀 STAGE 1: Starting rule-based header detection for '{sheet_name}'")
            logger.warning(f"📊 STAGE 1: Processing {len(df)} rows using identified columns")
            
            logger.warning(f"📊 STAGE 1: Creating chunks and detecting headers...")
            chunks, header_registry = stage1.create_sequential_chunks_enhanced(df, desc_col, unit_col, quantity_col)
            
            if not chunks:
                logger.warning(f"❌ STAGE 1: No chunks created for {sheet_name}")
                all_results[sheet_name] = {"error": f"No processable chunks found in {sheet_name}"}
                continue
            
            if not header_registry:
                logger.warning(f"❌ STAGE 1: No headers registered for {sheet_name}")
                all_results[sheet_name] = {"error": f"No headers found in {sheet_name}"}
                continue
            
            logger.warning(f"✅ STAGE 1 COMPLETE: Created {len(chunks)} chunks and registered {len(header_registry)} headers for '{sheet_name}'")
            
            # STAGE 2: AI-powered header verification
            logger.warning(f"🤖 STAGE 2: Starting AI header verification for '{sheet_name}'")
            logger.warning(f"🤖 STAGE 2: Processing {len(header_registry)} headers for AI verification...")
            
            # DEBUG: Check what type header_registry is
            logger.info(f"🔍 DEBUG: header_registry type = {type(header_registry)}, count = {len(header_registry)}")
            
            # Convert dictionary to list of values for Stage 2 processing
            if isinstance(header_registry, dict):
                header_list = list(header_registry.values())
                logger.info(f"🔍 DEBUG: Converted dict to list with {len(header_list)} items")
            else:
                header_list = header_registry
                logger.info(f"🔍 DEBUG: Using header_registry directly as list")
            
            # DEBUG: Check first item in header_list
            if header_list:
                first_item = header_list[0]
                logger.info(f"🔍 DEBUG: First header item type = {type(first_item)}")
            
            for i, header in enumerate(header_list, 1):
                try:
                    if isinstance(header, dict):
                        header_text = header.get('header_text', 'Unknown')
                        confidence = header.get('confidence', 0.0)
                    else:
                        header_text = getattr(header, 'header_text', f'Unknown_Object_{type(header)}')
                        confidence = getattr(header, 'confidence', 0.0)
                    
                    logger.info(f"🤖 STAGE 2: Header {i}/{len(header_list)}: '{header_text[:50]}...' (confidence: {confidence:.3f})")
                except Exception as e:
                    logger.error(f"❌ Error processing header {i}: {e}, header type: {type(header)}, content: {header}")
            
            verified_headers = await stage2.verify_headers(header_list)
            
            if not verified_headers:
                logger.warning(f"❌ STAGE 2: No verified headers for '{sheet_name}' - all headers rejected by AI")
                all_results[sheet_name] = {"error": f"No verified headers found in {sheet_name}"}
                continue
            
            logger.warning(f"✅ STAGE 2 COMPLETE: Verified {len(verified_headers)}/{len(header_registry)} headers for '{sheet_name}'")
            
            # Log details of verified headers
            for verified in verified_headers:
                if isinstance(verified, dict):
                    header_text = verified.get('header_text', 'Unknown')[:40]
                    category = verified.get('final_category', 'Unknown')
                    material = verified.get('final_material', 'Unknown')
                else:
                    header_text = getattr(verified, 'header_text', 'Unknown')[:40]
                    category = getattr(verified, 'final_category', 'Unknown')
                    material = getattr(verified, 'final_material', 'Unknown')
                
                logger.info(f"   ✅ Verified: '{header_text}...' -> Category: {category}, Material: {material}")
            
            # FIXED: Debug chunk ID mapping between Stage 2 and Stage 3
            logger.info(f"🔍 STAGE 2->3 MAPPING DEBUG:")
            
            # Handle both dict and object format for verified headers
            verified_header_map = {}
            for vh in verified_headers:
                if isinstance(vh, dict):
                    chunk_id = vh.get('chunk_id')
                else:
                    chunk_id = getattr(vh, 'chunk_id', None)
                
                if chunk_id is not None:
                    verified_header_map[chunk_id] = vh
            
            chunk_ids_from_stage1 = [chunk.chunk_id for chunk in chunks]
            chunk_ids_from_stage2 = list(verified_header_map.keys())
            
            logger.info(f"  - Chunk IDs from Stage 1: {chunk_ids_from_stage1}")
            logger.info(f"  - Chunk IDs from Stage 2: {chunk_ids_from_stage2}")
            
            missing_mappings = [cid for cid in chunk_ids_from_stage1 if cid not in verified_header_map]
            if missing_mappings:
                logger.warning(f"⚠️ Missing header context for chunks: {missing_mappings}")
            
            # STAGE 3: AI context inheritance for line items
            logger.warning(f"🧠 STAGE 3: Starting AI context inheritance for '{sheet_name}'")
            logger.warning(f"🧠 STAGE 3: Processing {len(chunks)} chunks with line item extraction")
            
            tasks = []
            processed_chunks = 0
            total_items = 0
            
            for i, chunk in enumerate(chunks, 1):
                header_context = verified_header_map.get(chunk.chunk_id)
                if header_context:
                    chunk_items_count = len(chunk.items)
                    total_items += chunk_items_count
                    logger.info(f"🧠 STAGE 3: Chunk {i}/{len(chunks)} - ID: {chunk.chunk_id}, Items: {chunk_items_count}")
                    
                    # Handle both dict and object format for header context
                    if isinstance(header_context, dict):
                        context_text = header_context.get('header_text', 'Unknown')[:40]
                    else:
                        context_text = getattr(header_context, 'header_text', 'Unknown')[:40]
                    
                    logger.info(f"   📋 Header context: '{context_text}...'")
                    
                    tasks.append(stage3.extract_chunk_details_enhanced(
                        chunk_info={'chunk_id': chunk.chunk_id}, 
                        verified_header=header_context, 
                        chunk_items=chunk.items
                    ))
                    processed_chunks += 1
                else:
                    logger.warning(f"⚠️ STAGE 3: Chunk {i}/{len(chunks)} - ID: {chunk.chunk_id} - NO HEADER CONTEXT, skipping")
            
            if not tasks:
                logger.warning(f"❌ STAGE 3: No tasks created for '{sheet_name}' - no chunks with valid header context")
                all_results[sheet_name] = {"error": f"No processable tasks found in {sheet_name}"}
                continue
            
            logger.info(f"🚀 STAGE 3: Executing {len(tasks)} parallel AI tasks for {processed_chunks} chunks ({total_items} total items)")
            
            # Execute all Stage 3 tasks in parallel
            logger.info(f"🧠 STAGE 3: Starting parallel AI processing...")
            line_item_results_list = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Flatten results and handle exceptions
            processed_items = []
            successful_tasks = 0
            failed_tasks = 0
            
            logger.info(f"🧠 STAGE 3: Processing {len(line_item_results_list)} task results...")
            
            for i, result in enumerate(line_item_results_list):
                if isinstance(result, Exception):
                    logger.error(f"❌ STAGE 3: Task {i+1}/{len(tasks)} FAILED: {result}")
                    failed_tasks += 1
                    continue
                
                task_items = len(result) if result else 0
                processed_items.extend(result)
                successful_tasks += 1
                logger.info(f"✅ STAGE 3: Task {i+1}/{len(tasks)} SUCCESS - extracted {task_items} items")
            
            logger.warning(f"✅ STAGE 3 COMPLETE: {successful_tasks}/{len(tasks)} tasks successful, {failed_tasks} failed")
            logger.warning(f"📊 Total items extracted: {len(processed_items)}")
            
            if not processed_items:
                logger.warning(f"❌ STAGE 3: No items processed for {sheet_name}")
                all_results[sheet_name] = {"error": f"No items successfully processed in {sheet_name}"}
                continue
            
            # Convert the final list of dataclass objects to a list of dictionaries for the frontend
            logger.info(f"📋 FINAL: Converting {len(processed_items)} items to JSON format for {sheet_name}")
            sheet_data_for_json = []
            
            for item in processed_items:
                try:
                    # Create a dictionary from the dataclass
                    item_dict = item.__dict__.copy()
                    
                    # Add the original row data from the DataFrame for the verification grid
                    if item.original_row_index < len(df):
                        original_row_data = df.iloc[item.original_row_index].to_dict()
                        item_dict['original_data'] = {k: v for k, v in original_row_data.items() if pd.notna(v)}
                    else:
                        item_dict['original_data'] = {}
                    
                    sheet_data_for_json.append(item_dict)
                    
                except Exception as e:
                    logger.error(f"❌ Error converting item to JSON: {e}")
                    continue

            all_results[sheet_name] = {
                "data": sheet_data_for_json,
                "stats": {
                    "total_items": len(processed_items),
                    "chunks_processed": processed_chunks,
                    "chunks_total": len(chunks),
                    "headers_verified": len(verified_headers),
                    "headers_registered": len(header_registry),
                    "description_column": desc_col,
                    "unit_column": unit_col,
                    "quantity_column": quantity_col,
                    "stage3_success_rate": f"{successful_tasks}/{len(tasks)}",
                    "column_mappings": column_mappings
                }
            }
            
            logger.info(f"🎉 SUCCESS: Sheet {sheet_name} completed - {len(processed_items)} items extracted through 3-stage pipeline")

    except Exception as e:
        logger.error(f"❌ ORCHESTRATOR ERROR: {e}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return {"error": str(e)}

    logger.info(f"🎉 ORCHESTRATOR COMPLETE: Processed {len(all_results)} sheets successfully")
    return all_results


def generate_summary_file(corrected_data: List[Dict]) -> BytesIO:
    """
    Takes corrected data and generates the final summary Excel file in memory.
    
    Args:
        corrected_data: List of dictionaries containing corrected BOQ data
        
    Returns:
        BytesIO object containing the Excel file
    """
    logger.info("📊 SUMMARIZER: Generating final summary file...")
    
    try:
        # Convert list of dictionaries to DataFrame
        data_df = pd.DataFrame(corrected_data)
        
        if data_df.empty:
            logger.warning("⚠️ No data provided for summary generation")
            # Create empty summary file
            data_df = pd.DataFrame({"Message": ["No data provided for summary"]})
        
        logger.info(f"📊 Generating summary for {len(data_df)} rows")
        
        # Generate material summary using our summarizer
        summary_df = generate_summary_dataframe(data_df)
        
        # Create an in-memory Excel file to send back to the user
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the material summary
            summary_df.to_excel(writer, sheet_name='Material_Summary', index=False)
            
            # Also include the original corrected data for reference
            if len(corrected_data) > 0 and "Message" not in data_df.columns:
                data_df.to_excel(writer, sheet_name='Corrected_Data', index=False)
        
        output.seek(0)
        
        logger.info("✅ Successfully generated summary Excel file")
        return output
        
    except Exception as e:
        logger.error(f"❌ Error generating summary file: {e}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        
        # Create error file
        error_output = BytesIO()
        error_df = pd.DataFrame({
            "Error": ["Failed to generate summary"],
            "Details": [str(e)],
            "Input_Rows": [len(corrected_data)]
        })
        
        with pd.ExcelWriter(error_output, engine='openpyxl') as writer:
            error_df.to_excel(writer, sheet_name='Error_Report', index=False)
        
        error_output.seek(0)
        return error_output


async def run_complete_boq_processing(file_path: str, selected_sheets: List[str], api_key: str, cached_analysis=None) -> Dict:
    """
    Complete BOQ processing pipeline - from file to extracted data.
    This is the main entry point for the orchestrator.
    
    Args:
        file_path: Path to the Excel file
        selected_sheets: List of worksheet names to process
        api_key: Gemini API key
        cached_analysis: Pre-computed analysis result (optional, avoids re-running analysis)
        
    Returns:
        Dictionary with extraction results and statistics
    """
    logger.info(f"🚀 COMPLETE BOQ PROCESSING: Starting pipeline for {len(selected_sheets)} sheets")
    
    try:
        # Run the extraction pipeline with cached analysis
        extraction_results = await run_extraction_pipeline(file_path, selected_sheets, api_key, cached_analysis)
        
        # Calculate overall statistics
        total_items = 0
        total_sheets_processed = 0
        total_errors = 0
        
        for sheet_name, result in extraction_results.items():
            if "error" in result:
                total_errors += 1
                logger.warning(f"❌ Sheet {sheet_name} failed: {result['error']}")
            else:
                total_sheets_processed += 1
                if "data" in result:
                    total_items += len(result["data"])
                elif isinstance(result, list):
                    total_items += len(result)
                logger.info(f"✅ Sheet {sheet_name} processed: {len(result.get('data', []))} items")
        
        success_rate = total_sheets_processed / len(selected_sheets) if selected_sheets else 0
        
        logger.info(f"🎉 PIPELINE COMPLETE: {total_sheets_processed}/{len(selected_sheets)} sheets successful ({success_rate:.1%})")
        logger.info(f"📊 TOTAL ITEMS EXTRACTED: {total_items}")
        
        return {
            "extraction_results": extraction_results,
            "overall_stats": {
                "total_sheets_requested": len(selected_sheets),
                "total_sheets_processed": total_sheets_processed,
                "total_items_extracted": total_items,
                "total_errors": total_errors,
                "processing_success_rate": success_rate
            }
        }
        
    except Exception as e:
        logger.error(f"❌ COMPLETE BOQ PROCESSING ERROR: {e}")
        return {
            "error": str(e),
            "overall_stats": {
                "total_sheets_requested": len(selected_sheets),
                "total_sheets_processed": 0,
                "total_items_extracted": 0,
                "total_errors": len(selected_sheets),
                "processing_success_rate": 0
            }
        }